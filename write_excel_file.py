from openpyxl import load_workbook


def preencher_planilhas(planilha_caminho, dados, codigo_referencia):
    """
    Preenche a planilha com os dados extraídos do PDF.

    Args:
        planilha_caminho (str): Caminho para o arquivo da planilha
        dados (dict): Dicionário com os dados extraídos do PDF
        codigo_referencia (str): Código de referência para buscar na planilha

    Returns:
        bool: True se a operação foi bem-sucedida, False caso contrário
    """
    print(f"Preenchendo planilha para código: {codigo_referencia}")

    try:
        # Abrir a planilha
        wb = load_workbook(planilha_caminho)
        atualizacoes_totais = 0

        # Verificar se os dados necessários existem
        if not dados or 'Nome_Solicitante' not in dados or 'CNPJ' not in dados:
            print(f"Dados incompletos para o código {codigo_referencia}")
            return False

        # Limpar e preparar os dados
        nome = dados.get('Nome_Solicitante', '').strip()
        cnpj = dados.get('CNPJ', '').strip()

        if not nome or not cnpj:
            print(f"Dados vazios para o código {codigo_referencia}")
            return False

        print(f"Dados a serem preenchidos: Nome={nome}, CNPJ={cnpj}")

        # Processar cada aba relevante
        for aba_nome in ['2025', '2024']:
            try:
                if aba_nome in wb.sheetnames:
                    ws = wb[aba_nome]
                    atualizacoes_aba = 0

                    # Determinar o intervalo de linhas a processar
                    linha_inicial = 6  # Começar da linha 6
                    linha_final = ws.max_row

                    print(f"Processando aba {aba_nome}, linhas {linha_inicial} até {linha_final}")

                    for linha in range(linha_inicial, linha_final + 1):
                        try:
                            # Obter o valor da célula na coluna B
                            valor_celula = ws[f'B{linha}'].value

                            # Comparar como strings, ignorando espaços extras
                            if valor_celula and str(valor_celula).strip() == str(codigo_referencia).strip():
                                print(f"Correspondência encontrada na linha {linha}")

                                # Preencher as células C e D
                                ws[f'C{linha}'].value = nome
                                ws[f'D{linha}'].value = cnpj

                                atualizacoes_aba += 1
                                atualizacoes_totais += 1
                                print(f"Célula atualizada: {aba_nome} linha {linha}")

                                # Não interromper o loop para processar todas as ocorrências
                                # Remova o comentário abaixo se quiser processar apenas a primeira ocorrência
                                # break
                        except Exception as e:
                            print(f"Erro ao processar linha {linha}: {e}")
                            continue

                    print(f"Total de atualizações na aba {aba_nome}: {atualizacoes_aba}")

                    # Salvar a cada aba processada para garantir que as alterações sejam mantidas
                    if atualizacoes_aba > 0:
                        wb.save(planilha_caminho)
                        print(f"Alterações na aba {aba_nome} salvas")
            except Exception as e:
                print(f"Erro ao processar aba {aba_nome}: {e}")
                continue

        # Salvamento final
        if atualizacoes_totais > 0:
            wb.save(planilha_caminho)
            print(f"Planilha salva com sucesso: {planilha_caminho}")
            print(f"Total de células atualizadas: {atualizacoes_totais}")
            return True
        else:
            print(f"Nenhuma célula foi atualizada para o código {codigo_referencia}")
            return False
    except Exception as e:
        print(f"Erro ao processar planilha: {e}")
        return False



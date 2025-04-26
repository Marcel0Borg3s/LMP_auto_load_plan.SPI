from openpyxl import load_workbook

def preencher_planilhas(planilha_caminho, dados, codigo_referencia):
    # Abrir a planilha
    wb = load_workbook(planilha_caminho)

    for aba_nome in ['2025', '2024']:
        if aba_nome in wb.sheetnames:
            ws = wb[aba_nome]

            for linha in range(6, ws.max_row + 1):  # Começar da linha 6
                valor_celula = ws[f'B{linha}'].value
                if valor_celula == codigo_referencia:
                    # Quando encontrar o código
                    nome = dados.get('Nome_Solicitante', '').replace('Solicitante/Endereço:', '').replace('Applicant/Address:', '').strip()
                    cnpj = dados.get('CNPJ', '')

                    ws[f'C{linha}'] = nome  # Coluna C
                    ws[f'D{linha}'] = cnpj  # Coluna D
                    break  # Já encontrou, pode sair do for
    # Salvar a planilha
    wb.save(planilha_caminho)



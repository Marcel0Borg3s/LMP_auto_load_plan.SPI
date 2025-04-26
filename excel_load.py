import os
from openpyxl import load_workbook

# planilha
pasta_base = r'E:\Programer\Python\Projetos\LMP\LMP_SPI_plan\Assets'
planilha_caminho = os.path.join(pasta_base, 'Processos - SPI.xlsm')

# 1. Abrir a planilha
wb = load_workbook(planilha_caminho)

# 2. Seleção de abas
abas_alvo = ['2025', '2024']

# 3. Para cada aba, procurar os codigos
codigos_encontrados = []

for aba_nome in abas_alvo:
    ws = wb[aba_nome]
    for linha in range(6, ws.max_row + 1):  # comecamos da linha 6 pra pular cabeçalho
        valor_celula = ws[f'B{linha}'].value
        if valor_celula and valor_celula.startswith('LMP'):
            codigos_encontrados.append(valor_celula)

# 4. Exibir o que foi encontrado
print("Códigos encontrados:")
for codigo in codigos_encontrados:
    print(f"- {codigo}")


